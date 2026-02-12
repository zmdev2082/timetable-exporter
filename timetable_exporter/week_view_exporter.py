from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time, timedelta
from typing import Any
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@dataclass
class WeekViewConfig:
    day_col: str
    start_time_col: str
    end_time_col: str | None
    duration_col: str | None
    summary_col: str
    week_pattern_col: str | None
    week_pattern_prefix: str | None
    week_pattern_full_term_tokens: list[str]
    week_pattern_full_term_label: str | None
    summary_transform: dict | None
    description_col: str | None
    summary_annotation: dict | None
    summary_format: str | None
    days: list[str]
    start_time: time
    end_time: time
    interval_minutes: int
    title: str | None
    include_week_pattern: bool
    footer: dict | None


DEFAULT_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def _parse_time(value: Any) -> time | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    parsed = pd.to_datetime(str(value), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().time()


def _normalize_day(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _to_minutes(t: time) -> int:
    return t.hour * 60 + t.minute


def _minutes_to_time(minutes: int) -> time:
    return time(hour=minutes // 60, minute=minutes % 60)


def _load_week_view_config(config: dict) -> WeekViewConfig:
    columns = config.get("columns", {})
    layout = config.get("layout", {})
    title = config.get("title")
    include_week_pattern = bool(config.get("include_week_pattern", True))
    summary_transform = config.get("summary_transform")
    description_col = columns.get("description")
    summary_annotation = config.get("summary_annotation")
    summary_format = config.get("summary_format")
    week_pattern_prefix = config.get("week_pattern_prefix")
    week_pattern_full_term_tokens = config.get("week_pattern_full_term_tokens") or []
    week_pattern_full_term_label = config.get("week_pattern_full_term_label")
    footer = config.get("footer")

    day_col = columns.get("day")
    start_time_col = columns.get("start_time")
    end_time_col = columns.get("end_time")
    duration_col = columns.get("duration")
    summary_col = columns.get("summary")
    week_pattern_col = columns.get("week_pattern")

    if not day_col or not start_time_col or not summary_col:
        raise ValueError("week view config requires columns.day, columns.start_time, columns.summary")

    days = layout.get("days") or DEFAULT_DAYS
    start_time = _parse_time(layout.get("start_time") or "08:00")
    end_time = _parse_time(layout.get("end_time") or "19:00")
    interval_minutes = int(layout.get("interval_minutes") or 60)

    if start_time is None or end_time is None:
        raise ValueError("week view config layout.start_time and layout.end_time must be valid times")

    return WeekViewConfig(
        day_col=day_col,
        start_time_col=start_time_col,
        end_time_col=end_time_col,
        duration_col=duration_col,
        summary_col=summary_col,
        week_pattern_col=week_pattern_col,
        week_pattern_prefix=week_pattern_prefix,
        summary_transform=summary_transform,
        description_col=description_col,
        summary_annotation=summary_annotation,
        summary_format=summary_format,
        week_pattern_full_term_tokens=week_pattern_full_term_tokens,
        week_pattern_full_term_label=week_pattern_full_term_label,
        days=days,
        start_time=start_time,
        end_time=end_time,
        interval_minutes=interval_minutes,
        title=title,
        include_week_pattern=include_week_pattern,
        footer=footer,
    )


def _extract_summary_annotation(row: pd.Series, cfg: WeekViewConfig) -> str | None:
    ann_cfg = cfg.summary_annotation or {}
    if not isinstance(ann_cfg, dict) or not ann_cfg:
        return None

    col = ann_cfg.get("column") or cfg.description_col
    if not col:
        return None

    raw = row.get(col)
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    text = str(raw)

    regex = ann_cfg.get("regex")
    if regex:
        try:
            m = re.search(str(regex), text)
        except re.error:
            return None
        if not m:
            return None
        group = ann_cfg.get("group", 1)
        try:
            value = m.group(int(group))
        except Exception:
            value = m.group(1)
        value = str(value).strip()
        return value or None

    return None


def _apply_cell_styles(cell, fill_color: str | None = None, bold: bool = False, align_center: bool = True):
    if fill_color:
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    if bold:
        cell.font = Font(bold=True)
    if align_center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _color_for_key(key: str, palette: list[str]) -> str | None:
    if not key or not palette:
        return None
    idx = abs(hash(key)) % len(palette)
    return palette[idx]


def _apply_summary_transform(value: Any, transform: dict | None) -> str:
    text = str(value).strip() if value is not None else ""
    if not transform or not text:
        return text

    split_on = transform.get("split_on")
    take = transform.get("take", 0)
    if split_on:
        if isinstance(split_on, str):
            split_on = [split_on]
        for sep in split_on:
            parts = text.split(sep)
            if parts:
                try:
                    text = parts[int(take)].strip()
                except Exception:
                    text = parts[0].strip()
    return text


def _normalize_week_pattern(value: Any, prefix: str | None, full_term_tokens: list[str], full_term_label: str | None) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None

    if prefix and not text.startswith(prefix):
        return None

    if text in full_term_tokens and full_term_label:
        return full_term_label

    wk_index = text.find("WK")
    if wk_index == -1:
        return None

    wk_text = text[wk_index:].strip()
    if "(" in wk_text:
        wk_text = wk_text.split("(", 1)[0].strip()
    return wk_text


def render_week_view_worksheet(ws, df: pd.DataFrame, config: dict) -> None:
    cfg = _load_week_view_config(config)

    palette = config.get("formatting", {}).get("palette", [])

    formatting = config.get("formatting", {})
    header_fill = formatting.get("header_fill", "D9D9D9")
    time_fill = formatting.get("time_fill", "F2F2F2")
    border_style = formatting.get("border", "thin")
    day_column_width = formatting.get("day_column_width", 15)
    time_column_width = formatting.get("time_column_width", 10)

    thin_side = Side(style=border_style)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    row_offset = 1
    col_offset = 2  # column A reserved for time labels

    if cfg.title:
        ws.cell(row=1, column=1, value=cfg.title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_offset + len(cfg.days) - 1)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_offset = 2

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = time_column_width
    for idx in range(len(cfg.days)):
        ws.column_dimensions[get_column_letter(col_offset + idx)].width = day_column_width

    # Day headers
    for idx, day in enumerate(cfg.days):
        cell = ws.cell(row=row_offset, column=col_offset + idx, value=day)
        _apply_cell_styles(cell, fill_color=header_fill, bold=True)
        cell.border = cell_border

    # Time slots
    start_minutes = _to_minutes(cfg.start_time)
    end_minutes = _to_minutes(cfg.end_time)
    slot_count = int((end_minutes - start_minutes) / cfg.interval_minutes)

    for i in range(slot_count):
        slot_start = start_minutes + i * cfg.interval_minutes
        label = _minutes_to_time(slot_start).strftime("%H:%M")
        cell = ws.cell(row=row_offset + 1 + i, column=1, value=label)
        _apply_cell_styles(cell, fill_color=time_fill, bold=True)
        cell.border = cell_border
        ws.row_dimensions[row_offset + 1 + i].height = config.get("formatting", {}).get("row_height", 22)

    # Apply borders to all day/time cells to avoid empty gaps
    for i in range(slot_count):
        for j in range(len(cfg.days)):
            cell = ws.cell(row=row_offset + 1 + i, column=col_offset + j)
            cell.border = cell_border

    # Aggregate bookings
    entries: dict[tuple, set[str]] = {}
    for _, row in df.iterrows():
        day = _normalize_day(row.get(cfg.day_col))
        if not day:
            continue
        if day not in cfg.days:
            continue

        start_t = _parse_time(row.get(cfg.start_time_col))
        if start_t is None:
            continue

        end_t = None
        if cfg.end_time_col:
            end_t = _parse_time(row.get(cfg.end_time_col))
        if end_t is None and cfg.duration_col:
            duration = row.get(cfg.duration_col)
            try:
                minutes = int(pd.to_timedelta(duration).total_seconds() / 60)
            except Exception:
                minutes = int(float(duration) * 60) if duration is not None else 0
            end_t = (datetime.combine(datetime.today(), start_t) + timedelta(minutes=minutes)).time()

        if end_t is None:
            continue

        base_summary = _apply_summary_transform(row.get(cfg.summary_col, ""), cfg.summary_transform)
        if not base_summary:
            continue

        annotation = _extract_summary_annotation(row, cfg)
        if annotation:
            if cfg.summary_format:
                summary = str(cfg.summary_format).format(summary=base_summary, annotation=annotation)
            else:
                summary = f"{base_summary} ({annotation})"
        else:
            summary = base_summary

        week_pattern = None
        if cfg.week_pattern_col:
            week_pattern = _normalize_week_pattern(
                row.get(cfg.week_pattern_col, ""),
                cfg.week_pattern_prefix,
                cfg.week_pattern_full_term_tokens,
                cfg.week_pattern_full_term_label,
            )
            if cfg.week_pattern_prefix and week_pattern is None:
                continue
        key = (day, start_t, end_t, summary)
        if key not in entries:
            entries[key] = set()
        if week_pattern:
            entries[key].add(week_pattern)

    # Render aggregated blocks
    occupied: dict[int, set[int]] = {col_offset + idx: set() for idx in range(len(cfg.days))}
    for (day, start_t, end_t, summary), patterns in entries.items():
        fill_color = _color_for_key(summary, palette)

        start_min = _to_minutes(start_t)
        end_min = _to_minutes(end_t)
        start_index = int((start_min - start_minutes) / cfg.interval_minutes)
        end_index = int((end_min - start_minutes) / cfg.interval_minutes)

        if start_index < 0 or end_index <= 0:
            continue
        end_index = max(end_index, start_index + 1)

        day_col_index = col_offset + cfg.days.index(day)
        block_start_row = row_offset + 1 + start_index
        block_end_row = row_offset + 1 + min(end_index, slot_count) - 1

        # Build display with merged week patterns
        if cfg.include_week_pattern and patterns:
            stripped = [p.replace("WK ", "", 1) if p.startswith("WK ") else p for p in patterns]

            def _sort_key(text: str) -> tuple:
                nums: list[int] = []
                for part in text.replace(" ", "").split(","):
                    if "-" in part:
                        start, end = part.split("-", 1)
                        if start.isdigit():
                            nums.append(int(start))
                        if end.isdigit():
                            nums.append(int(end))
                    elif part.isdigit():
                        nums.append(int(part))
                return (min(nums) if nums else 9999, text)

            stripped = sorted(stripped, key=_sort_key)
            merged = "WK " + ", ".join(stripped)
            display = f"{summary}\n({merged})"
        else:
            display = summary

        # Check if any cell in the block already has conflicting content or overlap
        has_conflict = False
        for r in range(block_start_row, block_end_row + 1):
            if r in occupied[day_col_index]:
                has_conflict = True
                break
            cell_value = ws.cell(row=r, column=day_col_index).value
            if cell_value and str(cell_value).strip() != display:
                has_conflict = True
                break

        target_cell = ws.cell(row=block_start_row, column=day_col_index)
        existing = str(target_cell.value).strip() if target_cell.value else ""
        if existing:
            if existing == display or display in existing.split("\n"):
                target_cell.value = existing
            else:
                target_cell.value = (existing + "\n" + display).strip()
        else:
            target_cell.value = display
        target_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if fill_color:
            for r in range(block_start_row, block_end_row + 1):
                ws.cell(row=r, column=day_col_index).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Merge blocks for multi-hour sessions when no conflicts exist
        if block_end_row > block_start_row and not has_conflict:
            ws.merge_cells(start_row=block_start_row, start_column=day_col_index, end_row=block_end_row, end_column=day_col_index)

        # Apply borders to block
        for r in range(block_start_row, block_end_row + 1):
            ws.cell(row=r, column=day_col_index).border = cell_border
            occupied[day_col_index].add(r)

    # Footer notes (simple): each line is a single merged cell (plain text only).
    if cfg.footer:
        footer_start = row_offset + 1 + slot_count + 2
        footer_lines = cfg.footer.get("lines", [])

        footer_col = col_offset
        footer_end_col = col_offset + len(cfg.days) - 1
        if footer_end_col < footer_col:
            footer_end_col = footer_col

        for idx, line in enumerate(footer_lines):
            display_text = (line.get("text") or "").strip()
            if not display_text:
                continue
            row = footer_start + idx

            cell = ws.cell(row=row, column=footer_col, value=display_text)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if footer_end_col > footer_col:
                ws.merge_cells(start_row=row, start_column=footer_col, end_row=row, end_column=footer_end_col)



def build_week_view_workbook(df: pd.DataFrame, config: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Schedule"
    render_week_view_worksheet(ws, df, config)
    return wb
